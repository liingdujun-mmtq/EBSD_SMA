import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib import cm
import tkinter as tk
from tkinter import scrolledtext
from tkinter import messagebox
from tkinter import END
from tkinter import filedialog
from tkinter.ttk import Progressbar
import os
import pickle
import time
from matplotlib.ticker import FormatStrFormatter

## Used function
def getR(phi,theta,psi):
    Z1=np.asmatrix([[np.cos(phi),np.sin(phi),0],
                   [-1*np.sin(phi),np.cos(phi),0],
                   [0,0,1]])
    N=np.asmatrix([[1,0,0],
                  [0,np.cos(theta),np.sin(theta)],
                  [0,-1*np.sin(theta),np.cos(theta)]])
    Z=np.asmatrix([[np.cos(psi),np.sin(phi),0],
                  [-1*np.sin(psi),np.cos(psi),0],
                  [0,0,1]])
    R=Z1*N*Z
    return np.array(R)

def get100(R):
    a1=R[0]
    a2=R[1]
    a3=R[2]
    a100=[a1,a2,a3]
    return a100

def get110(a):
    a110=[]
    for i in range(len(a)-1):
        for j in range(i+1,len(a)):
            a110.append((a[i]+a[j])/2)
            a110.append((a[i]-a[j])/2)
    return a110

def get111(a):
    a1=a[0]+a[1]+a[2]
    a2=-a[0]+a[1]+a[2]
    a3=a[0]-a[1]+a[2]
    a4=a[0]+a[1]-a[2]
    a111=[a1,a2,a3,a4]
    return a111

def eulor_to_OR(beta):
    beta_rad=np.deg2rad(beta)
    R=getR(beta_rad[0],beta_rad[1],beta_rad[2])
    a_100=get100(R)
    a_110=get110(a_100)
    a_111=get111(a_100)
    a=[a_100,a_110,a_111]
    return a

def getangle(a1,a2):
    angle_rad=np.arccos(np.dot(a1,a2)/(np.linalg.norm(a1)*np.linalg.norm(a2)))
    angle=np.rad2deg(angle_rad)
    return angle

def get_SMA(aSMA1,aSMA2):
    All_angle=[]
    for i in aSMA1:
        for j in aSMA2:
            All_angle.append(getangle(i,j))
            All_angle.append(getangle(i,-j))
    return np.min(All_angle)

## Data read function
def readxlsx(xlsx_path):
    workbook=openpyxl.load_workbook(xlsx_path)
    sheet= workbook.worksheets[0]
    maxcol=sheet.max_column
    mincol=sheet.min_column 
    minrow=sheet.min_row 
    maxrow=sheet.max_row
    global data
    data=sheet.iter_rows(min_row=2,values_only=True)
    return data

def readdata(data):
    x=[]
    y=[]
    eulor_angle=[]
    BS=[]
    for i in data:
        i=list(i)
        x.append(i[0])
        y.append(i[1])
        eulor_angle.append(i[2:5])
        BS.append(i[5])
    return([x,y,eulor_angle,BS])

# get max x and y
def get_pixlen(A):
    Alen=0
    for i in A:
        if i == 0:
            Alen=Alen+1
    return Alen

def getL(x,y):
    xlen=get_pixlen(y)
    ylen=get_pixlen(x)
    L=(xlen+ylen)*xlen*x[1]*1e-6
    return ([L,xlen,ylen])
    #with open('ScanL.txt', 'w') as f:
    #    f.write("%f\n" %L)

## Show Scale Marker
def plotscalemarker(x,y,color="tab:red"):
    markersize=[100,50,20,10,5,2,1]
    for i in markersize:
        if i<0.2*np.max(x):
            plt.plot([0+0.02*np.max(x),i+0.02*np.max(x)],[0.98*np.max(y),0.98*np.max(y)],color=color,linewidth=3)
            plt.text(0+0.02*np.max(x),0.96*np.max(y),"%s μm" %(i),fontsize=16,color=color)
            break


## Show BC image
def plotBSimage(x,y,BS,cmap="Greys"):
    
    plt.scatter(x,y,c=BS,cmap=cmap)
    plt.xlim(np.min(x),np.max(x))
    plt.ylim(np.max(y),np.min(y))
    plt.xticks([])
    plt.yticks([])
    plt.tight_layout()
    #plt.savefig("100-SMA.png",dpi=300)
    
def plotSMAimage(SMA,xSMA,ySMA,color="tab:red",minang=10,alpha=1,edgecolors="none"):
        
        SMAplot1,SMAplotx1,SMAploty1=highSMA(SMA,xSMA,ySMA,minang)
        plt.scatter(SMAplotx1,SMAploty1,c=color,s=4,alpha=alpha,edgecolors=edgecolors)
        plt.xlim(0,np.max(x))
        plt.ylim(np.max(y),0)
        plt.xticks([])
        plt.yticks([])
        plt.tight_layout()


## Calculation SMA
def outputSMA(x,y,eulor_angle):
    x2d=np.array(x).reshape(-1,xlen)
    y2d=np.array(y).reshape(-1,xlen)
    eulor_angle2d=np.array(eulor_angle).reshape(-1,xlen,3)
    xSMA=[]
    ySMA=[]
    alphaSMA100=[]
    alphaSMA110=[]
    alphaSMA111=[]

    for i in range(xlen-1):
        for j in range(ylen-1):
            a_g1=eulor_to_OR(eulor_angle2d[j][i])
      
            xSMA.append((x2d[j][i]+x2d[j+1][i])/2)
            ySMA.append((y2d[j][i]+y2d[j+1][i])/2)
        
            a_g2=eulor_to_OR(eulor_angle2d[j+1][i])
            alphaSMA100.append(get_SMA(a_g1[0],a_g2[0]))
            alphaSMA110.append(get_SMA(a_g1[1],a_g2[1]))
            alphaSMA111.append(get_SMA(a_g1[2],a_g2[2]))

            xSMA.append((x2d[j][i]+x2d[j][i+1])/2)
            ySMA.append((y2d[j][i]+y2d[j][i+1])/2)
            a_g2=eulor_to_OR(eulor_angle2d[j][i+1])
            alphaSMA100.append(get_SMA(a_g1[0],a_g2[0]))
            alphaSMA110.append(get_SMA(a_g1[1],a_g2[1]))
            alphaSMA111.append(get_SMA(a_g1[2],a_g2[2]))
        if i%10==0:
            print("%d/%d" %(i,xlen))
    print("Done!")
    return([alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA])

## Filter specified SMA
def highSMA(SMA,xSMA,ySMA,a):
    SMAplot=[]
    SMAplotx=[]
    SMAploty=[]
    for i in range(len(SMA)):
        if SMA[i] >=a:
            SMAplot.append(SMA[i])
            SMAplotx.append(xSMA[i])
            SMAploty.append(ySMA[i])
    return [SMAplot,SMAplotx,SMAploty]

## Filter angle
def filterAng(A,min_A):
    return([i for i in A if i>=min_A])


##########################################################################################
#GUI code


dataready=False
def main_window():
    root=tk.Tk()
    root.title('SMA Calculator Alpha')

    main_frame=tk.Frame(root)
    main_frame.pack(anchor='w')

    ### Read Data GUI
    Data_path=tk.StringVar(root,value="")
    tk.Label(main_frame,text="Data Files:",width=16).grid(row=0,column=0)
    tk.Entry(main_frame,textvariable=Data_path,state='readonly',width=50,justify='left').grid(row = 0, column = 1)

    def select_data_file():
        Data_path_=filedialog.askopenfilename()
        Data_path.set(Data_path_)

    tk.Button(main_frame, text = "Select Data Files", command = select_data_file,width=15).grid(row = 0, column = 2)

    ### config GUI
    cal_frame=tk.Frame(root)
    cal_frame.pack(anchor='w')

    tk.Label(cal_frame,text="Config:",width=16).grid(row=1,column=0)
    config_BS_var=tk.BooleanVar()
    config_BS=tk.Checkbutton(cal_frame,text="BS Data",variable=config_BS_var).grid(row=1,column=1)

    config_SMA_var=tk.BooleanVar()
    config_SMA=tk.Checkbutton(cal_frame,text="SMA Data",variable=config_SMA_var).grid(row=1,column=2)

    config_SMA_dis_var=tk.BooleanVar()
    config_SMA_dis=tk.Checkbutton(cal_frame,text="SMA distribution",variable=config_SMA_dis_var).grid(row=1,column=3)
    


    ### plot GUI
    plot_frame=tk.Frame(root)
    plot_frame.pack(anchor='w')
    tk.Label(plot_frame,text="Image:",width=16).grid(row=1,column=0)
    config_BS_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="BS Image",variable=config_BS_img).grid(row=1,column=1)

    config_SMA_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="SMA Image",variable=config_SMA_img).grid(row=1,column=2)

    config_SMABC_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="SMA+BC Image",variable=config_SMABC_img).grid(row=1,column=3)

    config_SMA_dis_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="SMA distribution",variable=config_SMA_dis_img).grid(row=1,column=4)

    config_save_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="Save image",variable=config_save_img).grid(row=2,column=1)
    config_notshow_img=tk.BooleanVar()
    tk.Checkbutton(plot_frame,text="Skip show image",variable=config_notshow_img).grid(row=2,column=2)

    ### custom GUI
    custom_frame=tk.Frame(root)
    custom_frame.pack(anchor='w')
    tk.Label(custom_frame,text="Custom:",width=16).grid(row=1,column=0)

    tk.Label(custom_frame,text="Custom F:",width=16).grid(row=2,column=0)
    custom_F=tk.Entry(custom_frame,textvariable=tk.IntVar(value=5))
    custom_F.grid(row=2,column=1)




    ### function GUI
    function_frame=tk.Frame(root)
    function_frame.pack(anchor='w')

    def Load_data():
        datapath=Data_path.get()
        if (os.path.exists(datapath)) and (datapath[-5::]==".xlsx"):
            data=readxlsx(datapath)
            global x,y,eulor_angle,BS
            x,y,eulor_angle,BS=readdata(data)
            global dataready
            dataready=True
            MessageBox.insert(END, "%s: Load Data: %s \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime()),datapath))
        else:
            messagebox.showerror(title="Error", message="Wrong Data File!")

    tk.Button(function_frame, text = "Read Data", command = Load_data,width=15).grid(row = 0, column = 1)

    def Cal():
        BS_config=config_BS_var.get()
        if BS_config:
            global BSdata
            BCdata=BS
            with open ("cache_BS.pickle", "wb") as file:
                pickle.dump((x,y,BS), file)
            MessageBox.insert(END, "%s BS Data Cached! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))

        
        SMA_config=config_SMA_var.get()
        if SMA_config:
            MessageBox.insert(END, "%s SMA Data Start! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))
            global L,xlen,ylen
            L,xlen,ylen=getL(x,y)

            global alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA
            alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA=outputSMA(x,y,eulor_angle)
            MessageBox.insert(END, "%s SMA Data Finished! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))
            with open ("cache_SMA.pickle", "wb") as file:
                pickle.dump((alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA,xlen,ylen), file)
            MessageBox.insert(END, "%s SMA Data Cached! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))
        #messagebox.showinfo(title="Info", message="Cal Finished!")

        SMA_distribution_config=config_SMA_dis_var.get()
        if SMA_distribution_config:
            global SMA_dis100,SMA_dis110,SMA_dis111
            SMA_dis100=filterAng(alphaSMA100,3)
            SMA_dis110=filterAng(alphaSMA110,3)
            SMA_dis111=filterAng(alphaSMA111,3)
            with open ("cache_SMA_dis.pickle", "wb") as file:
                pickle.dump((SMA_dis100,SMA_dis110,SMA_dis111,L), file)


    def Configcal():
        if dataready:
            Cal()
        else:
            messagebox.showerror(title="Error", message="Wrong Data File!")
    
    tk.Button(function_frame, text = "Start Cal", command = Configcal,width=15).grid(row = 0, column = 3)

    def Showimage():
        BS_img=config_BS_img.get()
        save_img=config_save_img.get()
        notshow_img=config_notshow_img.get()
        if BS_img:
            plt.figure(num="BS Image",figsize=(6,6))
            plotBSimage(x,y,BS,cmap="Greys")
            plotscalemarker(x,y,color="tab:red")
            if save_img:
                plt.savefig("img/BS.png",dpi=300)
        
        SMA_img=config_SMA_img.get()
        if SMA_img:
            plt.figure(num="100-SMA Image",figsize=(6,6))
            plotSMAimage(alphaSMA100,xSMA,ySMA,color="tab:red",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/100-SMA.png",dpi=300)

            plt.figure(num="110-SMA Image",figsize=(6,6))
            plotSMAimage(alphaSMA110,xSMA,ySMA,color="tab:green",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/110-SMA.png",dpi=300)

            plt.figure(num="111-SMA Image",figsize=(6,6))
            plotSMAimage(alphaSMA111,xSMA,ySMA,color="tab:blue",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/111-SMA.png",dpi=300)
            plt.figure(num="All SMA Image",figsize=(6,6))
            plotSMAimage(alphaSMA100,xSMA,ySMA,color="tab:red",minang=10,alpha=0.3,edgecolors="none")
            plotSMAimage(alphaSMA110,xSMA,ySMA,color="tab:green",minang=10,alpha=0.3,edgecolors="none")
            plotSMAimage(alphaSMA111,xSMA,ySMA,color="tab:blue",minang=10,alpha=0.3,edgecolors="none")
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/ALL-SMA.png",dpi=300)

        SMABC_img=config_SMABC_img.get()
        if SMABC_img:
            plt.figure(num="100-SMA+BS Image",figsize=(6,6))
            plotBSimage(x,y,BS,cmap="Greys")
            plotSMAimage(alphaSMA100,xSMA,ySMA,color="tab:red",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/100-SMA+BS.png",dpi=300)

            plt.figure(num="110-SMA+BS Image",figsize=(6,6))
            plotBSimage(x,y,BS,cmap="Greys")
            plotSMAimage(alphaSMA110,xSMA,ySMA,color="tab:green",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/110-SMA+BS.png",dpi=300)

            plt.figure(num="111-SMA+BS Image",figsize=(6,6))
            plotBSimage(x,y,BS,cmap="Greys")
            plotSMAimage(alphaSMA111,xSMA,ySMA,color="tab:blue",minang=10)
            plotscalemarker(x,y,color="black")
            if save_img:
                plt.savefig("img/111-SMA+BS.png",dpi=300)
        SMA_dis_img=config_SMA_dis_img.get()
        if SMA_dis_img:
            plt.figure(num="SMA-100 distribution",figsize=(6,6))
            count=len(SMA_dis100)
            #F=5
            F=int(custom_F.get())
            plt.hist(SMA_dis100,bins=range(3,45),weights=[1/L/10**F]*count,color="tab:red",alpha=0.8)
            plt.xticks(range(0,46,5),fontsize=16)
            plt.yticks(fontsize=16)
            plt.gca().yaxis.set_major_formatter(FormatStrFormatter('%1.1f'))
            plt.xlabel("100-SMA (°)",fontsize=20)
            plt.ylabel("SMA Density ($×10^{%d}$ m$^{-1}$)" %F,fontsize=20)
            plt.tight_layout()
            if save_img:
                plt.savefig("img/100-SMA+dis.png",dpi=300)
            MessageBox.insert(END, "100-SMA Mean: %4.2f ± %4.2f \n" % (np.mean(SMA_dis100),np.std(SMA_dis100)) )
            MessageBox.insert(END, "Total 100-SMA density = %4.2e \n" %(len(SMA_dis100)/L))
            
            
            plt.figure(num="SMA-110 distribution",figsize=(6,6))
            count=len(SMA_dis110)
            plt.hist(SMA_dis110,bins=range(3,45),weights=[1/L/10**F]*count,color="tab:green",alpha=0.8)
            plt.xticks(range(0,46,5),fontsize=16)
            plt.yticks(fontsize=16)
            plt.gca().yaxis.set_major_formatter(FormatStrFormatter('%1.1f'))
            plt.xlabel("110-SMA (°)",fontsize=20)
            plt.ylabel("SMA Density ($×10^{%d}$ m$^{-1}$)" %F,fontsize=20)
            plt.tight_layout()
            if save_img:
                plt.savefig("img/110-SMA+dis.png",dpi=300)
            MessageBox.insert(END, "110-SMA Mean: %4.2f ± %4.2f \n" % (np.mean(SMA_dis110),np.std(SMA_dis110)) )
            MessageBox.insert(END, "Total 110-SMA density = %4.2e \n" %(len(SMA_dis110)/L))

            plt.figure(num="SMA-111 distribution",figsize=(6,6))
            count=len(SMA_dis111)
            plt.hist(SMA_dis111,bins=range(3,45),weights=[1/L/10**F]*count,color="tab:blue",alpha=0.8)
            plt.xticks(range(0,46,5),fontsize=16)
            plt.yticks(fontsize=16)
            plt.gca().yaxis.set_major_formatter(FormatStrFormatter('%1.1f'))
            plt.xlabel("111-SMA (°)",fontsize=20)
            plt.ylabel("SMA Density ($×10^{%d}$ m$^{-1}$)" %F,fontsize=20)
            plt.tight_layout()
            if save_img:
                plt.savefig("img/111-SMA+dis.png",dpi=300)
            MessageBox.insert(END, "111-SMA Mean: %4.2f ± %4.2f \n" % (np.mean(SMA_dis111),np.std(SMA_dis111)) )
            MessageBox.insert(END, "Total 111-SMA density = %4.2e \n" %(len(SMA_dis111)/L))



        MessageBox.insert(END, "%s Plot! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))
        if not notshow_img:
            plt.show()
        else:
            plt.close('all')
        


    tk.Button(function_frame, text = "Show Image", command = Showimage,width=15).grid(row = 0, column = 5)

    def readcacheddata():
        if os.path.exists("cache_SMA.pickle"):
            global alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA,L,xlen,ylen
            with open("cache_SMA.pickle", "rb") as file:
                alphaSMA100,alphaSMA110,alphaSMA111,xSMA,ySMA,xlen,ylen=pickle.load(file)
        if os.path.exists("cache_BS.pickle"):
            global x,y,BS
            with open("cache_BS.pickle", "rb") as file:
                x,y,BS=pickle.load(file)
        if os.path.exists("cache_SMA_dis.pickle"):
            global SMA_dis100,SMA_dis110,SMA_dis111,L
            with open("cache_SMA_dis.pickle", "rb") as file:
                SMA_dis100,SMA_dis110,SMA_dis111,L=pickle.load(file)  

     
        MessageBox.insert(END, "%s Load Data from Cache! \n" %(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))



    tk.Button(function_frame, text = "Read Cached Data", command = readcacheddata,width=15).grid(row = 0, column = 7)


    ### log GUI
    log_frame=tk.Frame(root)
    log_frame.pack(anchor='w')
    MessageBox=scrolledtext.ScrolledText(log_frame, width=80,height=40)
    MessageBox.grid(row=1,column=0,columnspan=3,sticky="w")
    MessageBox.insert(END, "SMA Calculator by LingDuJun \n")
    MessageBox.insert(END, "================================================= \n")
    

    root.mainloop()

main_window()
